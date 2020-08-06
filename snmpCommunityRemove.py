import datetime
import time
import wexpect
import getpass
import openpyxl

def accessJumpBox(username, password):

    print('\n--- Attempting connection to ' + 'IS6 Server... ')
    ssh_newkey = 'Are you sure you want to continue connecting'
    session = wexpect.spawn('ssh ' + username + '@is6.hsnet.ufl.edu')

    idx = session.expect([ssh_newkey, 'word', wexpect.EOF])

    if idx == 0:
        session.sendline('yes')
        idx = session.expect([ssh_newkey, 'word', wexpect.EOF])

        if idx == 0:
            session.sendline(password)
    elif idx == 1:
        session.sendline(password)

    idx = session.expect(['$', wexpect.EOF])

    if idx == 0:
        print("--- Successful Login to JumpBox")
        return session
    else:
        print("--- Terminated program")
        exit()

def accessSwitches(session, switch, username, password):

    ssh_newkey = 'Are you sure you want to continue'
    telnet = 'Connection refused'
    
    session.sendline('ssh ' + switch[1])

    print('\n------------------------------------------------------')
    print('--- Attempting connection to: ' + switch[1])
    print('------------------------------------------------------\n')

    idx = session.expect([ssh_newkey, 'word', telnet, wexpect.EOF])

    if idx == 0:
        session.sendline('yes')
        time.sleep(.5)
        session.sendline(password)
    elif idx == 1:
        time.sleep(.5)
        session.sendline(password)
    elif idx == 2:
        session.sendline('telnet ' + switch[1])

        idx = session.expect(['name', wexpect.EOF])

        if idx == 0:
            session.sendline(username)
            idx = session.expect(['word', wexpect.EOF])
            session.sendline(password)

        else:
            print('Something''s wrong!')
            print('--- Terminated Program!!')
            exit()

    idx = session.expect(['>', '#', wexpect.EOF])
    
    print('--- Success Login to: ', switch[1])

    if idx == 0:
        session.sendline('en')
        idx = session.expect(['word:', wexpect.EOF])
        
    if idx == 0:
        session.sendline(password)
        session.expect(['#', wexpect.EOF])

    return session

def loadExcel():

    wb = openpyxl.load_workbook(r'P:\Script\SNMP-Community\SNMP_Community.xlsx')
    ws = wb.active

    excelData = list()
    
    for i in range(5, ws.max_row + 1):
        device = list()
        for j in range(1, ws.max_column):
            cellValue = ws.cell(row = i, column = j).value
            
            if cellValue == None:
                snmpCommunity = ws.cell(row = i, column = j + 2).value
                excelData[-1].append(snmpCommunity)
                break
            else:
                device.append(cellValue)
        if len(device) != 0:
            excelData.append(device)
    
    return excelData

def commandExecute(session, device):

    command = ''

    session.sendline('conf t')
    session.expect(['#', wexpect.EOF])
    
    for i in range(0, len(device[2:])):

        command = 'no ' + device[2:][i]
        session.sendline(command)
        session.expect(['#', wexpect.EOF])

    session.sendline('end')
    session.expect(['#', wexpect.EOF]) 

    session.sendline('wr')

    if device[0].startswith('Y') or device[0].startswith('G'):
        session.expect(['complete.', 'OK', wexpect.EOF], timeout=60)
    else:
        session.expect(['#', wexpect.EOF], timeout=60)
    

if __name__ == '__main__':

    cellNumber = 5
    print()
    print('+-------------------------------------------------------------+')
    print('|    Cisco L2 switches Interface Update tool...               |')
    print('|    Updated by a premade excel spread sheet                  |')
    print('|    Version 1.0.0                                            |')
    print('|    Compatible with C35xx, C37xx, C38xx, C65XX               |')
    print('|    Nexus 5K, 7K, 9K                                         |')
    print('|    Scripted by Ethan Park, July. 2020                       |')
    print('+-------------------------------------------------------------+')
    print()
    devices = loadExcel()
    print(devices)
    
    username = input("Enter your admin ID ==> ")
    password = getpass.getpass("Enter your password ==> ")
    print()
    
    for elem in devices:
        session = accessJumpBox(username, password)
        session = accessSwitches(session, elem, username, password)

        commandExecute(session, elem)
        print(session.before)
        
        session.close()
    