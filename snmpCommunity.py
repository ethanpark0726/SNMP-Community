import datetime
import time
import wexpect
import pprint
import getpass
import openpyxl
import parse
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill

def createExcelFile():
    
    # Excel File Creation
    nowDate = 'Report Date: ' + str(datetime.datetime.now().strftime('%Y-%m-%d'))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'SNMP Community'
    
    # Pretty display for the File
    font = Font(bold=True)
    alignment = Alignment(horizontal='center')
    bgColor = PatternFill(fgColor='BFBFBFBF', patternType='solid')
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    ws['A2'] = nowDate
    ws['A4'] = 'Hostname'
    ws['A4'].alignment = alignment
    ws['A4'].font = font
    ws['A4'].fill = bgColor
    ws['A4'].border = border

    ws['B4'] = 'IP Address'
    ws['B4'].alignment = alignment
    ws['B4'].font = font
    ws['B4'].fill = bgColor
    ws['B4'].border = border

    ws['C4'] = 'SNMP Community'
    ws['C4'].alignment = alignment
    ws['C4'].font = font  
    ws['C4'].fill = bgColor
    ws['C4'].border = border

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 60
    
    fileName = 'SNMP_Community.xlsx'
    wb.save(fileName)
    wb.close()

def saveExcelFile(deviceList, snmpCommunityList, cellNumber):

    fileName = 'SNMP_Community.xlsx'
    wb = openpyxl.load_workbook(fileName)
    ws = wb.active
    alignment = Alignment(horizontal='center')
    border = Border(left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))

    ws['A' + str(cellNumber)] = deviceList[0]
    ws['A' + str(cellNumber)].alignment = alignment
    ws['A' + str(cellNumber)].border = border

    ws['B' + str(cellNumber)] = deviceList[2]
    ws['B' + str(cellNumber)].alignment = alignment
    ws['B' + str(cellNumber)].border = border

    for i in range(len(snmpCommunityList)):
        ws['C' + str(i + cellNumber)] = snmpCommunityList[i]
        ws['C' + str(i + cellNumber)].alignment = alignment
        ws['C' + str(i + cellNumber)].border = border

    wb.save('SNMP_Community.xlsx')
    wb.close()

def accessJumpBox(username, password):

    print('\n--- Attempting connection to ' + 'IS6 Server... ')
    ssh_newkey = 'Are you sure you want to continue connecting'
    session = wexpect.spawn('ssh ' + username + '@yourServerIP')

    idx = session.expect([ssh_newkey, 'word', wexpect.EOF])

    if idx == 0:
        session.sendline('yes')
        idx = session.expect([ssh_newkey, 'word', wexpect.EOF])

        if idx == 0:
            session.sendline(password)
    elif idx == 1:
        session.sendline(password)

    idx = session.expect(['$', wexpect.EOF])

    if idx == 0:
        print("--- Successful Login to JumpBox")
        return session
    else:
        print("--- Terminated program")
        exit()

def accessSwitches(session, switch, username, password):

    if 'SSH' in switch[3]:
        ssh_newkey = 'Are you sure you want to continue'
        session.sendline('ssh ' + switch[2])

        print('\n------------------------------------------------------')
        print('--- Attempting connection to: ' + switch[2])
        print('------------------------------------------------------\n')

        idx = session.expect([ssh_newkey, 'word', wexpect.EOF])

        if idx == 0:
            session.sendline('yes')
            time.sleep(.5)
            session.sendline(password)
        elif idx == 1:
            session.sendline(password)
        
    else:
        session.sendline('telnet ' + switch[2])
        
        print('\n------------------------------------------------------')
        print('--- Attempting connection to: ' + switch[2])
        print('------------------------------------------------------\n')

        idx = session.expect(['name', wexpect.EOF])

        if idx == 0:
            session.sendline(username)
            idx = session.expect(['word', wexpect.EOF])
            session.sendline(password)

        else:
            print('Something''s wrong!')
            print('--- Terminated Program!!')
            exit()
    idx = session.expect(['>', '#', wexpect.EOF])
    print('--- Success Login to: ', switch[2])
 
    if idx == 0:
        session.sendline('en')
        idx = session.expect(['word:', wexpect.EOF])
        
    if idx == 0:
        session.sendline(password)
        session.expect(['#', wexpect.EOF])
    
    return session

def getDeviceList():
    deviceList = []

    file = open('0728.txt', 'r')

    for line in file:
        temp = line.split('\t')
        temp[-1] = temp[-1].replace('\n', '')
        deviceList.append(temp)
    file.close()

    return deviceList

def commandExecute(session, os):

    command = ''

    session.sendline('term length 0')
    session.expect(['#', wexpect.EOF])

    if os == 'IOS':
        command += 'sh run | i snmp-server com'
    elif os == 'NXOS':
        command += 'sh run | i \'snmp-server com\''
        
    session.sendline(command)
    session.expect(['#', wexpect.EOF])

    return session.before.splitlines()

if __name__ == '__main__':

    cellNumber = 5
    print()
    print('+-------------------------------------------------------------+')
    print('|    Cisco L2 switches SNMP Community Gathernig tool...       |')
    print('|    Version 1.0.0                                            |')
    print('|    Compatible with C35xx, C37xx, C38xx, C65XX               |')
    print('|    Nexus 5K, 7K, 9K                                         |')
    print('|    Scripted by Ethan Park, June. 2020                       |')
    print('+-------------------------------------------------------------+')
    print()
    username = input("Enter your admin ID ==> ")
    password = getpass.getpass("Enter your password ==> ")
    print()

    switchList = getDeviceList()
    createExcelFile()

    for elem in switchList:
        
        session = accessJumpBox(username, password)
        session = accessSwitches(session, elem, username, password)
        data = commandExecute(session, elem[1])
        switch = parse.Parse(data)
        saveExcelFile(elem, switch.getSNMPCommunity(), cellNumber)

        cellNumber += len(switch.getSNMPCommunity())
        session.close()
